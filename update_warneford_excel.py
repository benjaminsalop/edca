"""
Update Warneford_Comparison Excel with latest EDCA run data.
"""

import shutil
import re
import pandas as pd
import openpyxl

# ── Constants ────────────────────────────────────────────────────────────────
GFA_PER_LEVEL = 5690
NUM_LEVELS = 3
TOTAL_GFA = GFA_PER_LEVEL * NUM_LEVELS  # 17070

SRC = "/Users/benjaminsalop/Desktop/Warneford_Comparison_new_inputs_excl_lateral.xlsx"
DST = "/Users/benjaminsalop/Desktop/Warneford_Comparison_updated_May19.xlsx"

CSV_PATHS = {
    "6.6": "/Users/benjaminsalop/Desktop/Oxford/Research/edca/outputs/edca_run_6.6/summary_assemblies_ranked_6.6.csv",
    "7.5": "/Users/benjaminsalop/Desktop/Oxford/Research/edca/outputs/edca_run_7.5/summary_assemblies_ranked_7.5.csv",
    "9.0": "/Users/benjaminsalop/Desktop/Oxford/Research/edca/outputs/edca_run_9.0/summary_assemblies_ranked_9.0.csv",
}

# Typology definitions: (typology_name, floor_type, column_family_type)
# column_family_type: "rc" = ec_gravity_column, "steel" = uk_h*, "timber" = non-rc non-steel, None = any
TYPOLOGIES = [
    ("RC Flat Slab",           "flat_slab",       "rc"),
    ("2Way RC Slab",           "two_way_slab",    "rc"),
    ("RC Waffle Slab",         "waffle_slab",     "rc"),
    ("PT Flat Slab",           "pt_slab",         None),
    ("Composite Steel frame",  "composite_deck",  "steel"),
    ("Steel with Hollow slabs","hollowcore",       "steel"),
    ("Steel with CLT Slabs",   "clt_floor",       "steel"),
    ("Glulam with CLT Slabs",  "clt_floor",       "timber"),
]


def get_col_family_filter(df, col_type):
    """Return the column_family value matching col_type in this specific CSV."""
    if col_type is None:
        return None
    families = df["column_family"].unique()
    if col_type == "rc":
        candidates = [f for f in families if "ec_gravity_column" in f.lower()]
    elif col_type == "steel":
        # Steel = starts with uk_h (CHS/RHS/SHS sections)
        candidates = [f for f in families if f.lower().startswith("uk_h")]
    elif col_type == "timber":
        # Timber = not RC and not steel
        candidates = [f for f in families
                      if "ec_gravity_column" not in f.lower() and not f.lower().startswith("uk_h")]
    else:
        candidates = []
    if not candidates:
        raise ValueError(f"No column family found for type {col_type!r} in CSV (families: {list(families)})")
    if len(candidates) > 1:
        print(f"  Note: Multiple {col_type} column families found: {candidates}, using {candidates[0]}")
    return candidates[0]


# ── Helper functions ─────────────────────────────────────────────────────────

def compute_excl_lateral_carbon(row):
    """Sum of floor+beam+secondary_beam+column carbon per m2."""
    return (row["floor_carbon_per_m2"] + row["beam_carbon_per_m2"]
            + row["secondary_beam_carbon_per_m2"] + row["column_carbon_per_m2"])


def compute_material_carbon(row, material):
    """Sum excl-lateral carbon for a given material."""
    components = ["floor", "beam", "sec_beam", "column"]
    # sec_beam uses sec_beam prefix
    total = 0.0
    for comp in components:
        col = f"{comp}_carbon_{material}_per_m2"
        total += row.get(col, 0.0) or 0.0
    return total


def compute_material_mass(row, material):
    """Sum excl-lateral mass for a given material (kg/m2)."""
    components = ["floor", "beam", "sec_beam", "column"]
    total = 0.0
    for comp in components:
        col = f"{comp}_{material}_mass_per_m2"
        total += row.get(col, 0.0) or 0.0
    return total


def compute_total_mass_excl_lateral(row):
    """Sum all material masses excl lateral (kg/m2)."""
    materials = ["concrete", "structural_steel", "rebar", "pt", "timber", "screed"]
    return sum(compute_material_mass(row, m) for m in materials)


def get_best_row(df, floor_type, col_family_exact):
    """Return the DataFrame row with lowest total_embodied_carbon_per_m2 matching criteria."""
    mask = df["floor_type"] == floor_type
    if col_family_exact:
        mask = mask & (df["column_family"] == col_family_exact)
    candidates = df[mask].sort_values("total_embodied_carbon_per_m2")
    if candidates.empty:
        raise ValueError(f"No rows found for floor_type={floor_type!r}, col_family={col_family_exact!r}")
    return candidates.iloc[0]


def build_row_data(row, span_label, typology, csv_fname):
    """Build a dict of all values for one Assembly Audit row."""
    excl_lat = compute_excl_lateral_carbon(row)
    total_incl = row["total_embodied_carbon_per_m2"]
    removed_lat = total_incl - excl_lat

    concrete_mass = compute_material_mass(row, "concrete")
    rebar_mass = compute_material_mass(row, "rebar")
    steel_mass = compute_material_mass(row, "structural_steel")
    pt_mass = compute_material_mass(row, "pt")
    timber_mass = compute_material_mass(row, "timber")
    # screed mass (floor only per instructions)
    screed_mass = row.get("floor_screed_mass_per_m2", 0.0) or 0.0

    def to_tonnes_3level(kg_m2):
        return kg_m2 * NUM_LEVELS * GFA_PER_LEVEL / 1000

    return {
        "span": span_label,
        "typology": typology,
        "rank": int(row["rank_carbon"]),
        "assembly_family": row.get("assembly_family", ""),
        "floor_variant": str(row["floor_variant"]) if pd.notna(row["floor_variant"]) else "NaN",
        "beam_variant": str(row["beam_variant"]) if pd.notna(row["beam_variant"]) else "NaN",
        "sec_beam_variant": str(row["secondary_beam_variant"]) if pd.notna(row["secondary_beam_variant"]) else "NaN",
        "column_variant": str(row["column_variant"]) if pd.notna(row["column_variant"]) else "NaN",
        "lateral_variant": str(row["lateral_variant"]) if pd.notna(row["lateral_variant"]) else "NaN",
        "excl_lat_carbon": excl_lat,
        "total_incl_carbon": total_incl,
        "removed_lat_carbon": removed_lat,
        "concrete_3level_t": to_tonnes_3level(concrete_mass),
        "rebar_3level_t": to_tonnes_3level(rebar_mass),
        "steel_pt_3level_t": to_tonnes_3level(steel_mass + pt_mass),
        "timber_3level_t": to_tonnes_3level(timber_mass),
        "screed_carbon": row.get("floor_carbon_screed_per_m2", 0.0) or 0.0,
        "source_csv": csv_fname,
        # Carbon breakdown per m2
        "concrete_carbon": compute_material_carbon(row, "concrete"),
        "rebar_carbon": compute_material_carbon(row, "rebar"),
        "steel_carbon": compute_material_carbon(row, "structural_steel"),
        "timber_carbon": compute_material_carbon(row, "timber"),
        # Mass per m2 for span sheets
        "total_mass_kg_m2": compute_total_mass_excl_lateral(row),
        "concrete_mass_kg_m2": concrete_mass,
        "rebar_mass_kg_m2": rebar_mass,
        "steel_pt_mass_kg_m2": steel_mass + pt_mass,
        "timber_mass_kg_m2": timber_mass,
        "screed_mass_kg_m2": screed_mass,
    }


# ── Load CSVs and extract best rows ──────────────────────────────────────────

print("Loading CSVs...")
dfs = {span: pd.read_csv(path) for span, path in CSV_PATHS.items()}

# Build full dataset: list of (span_label, typology, row_data)
data = {}  # key: (span, typology) -> row_data

span_map = {"6.6": "6.6x6.6", "7.5": "6.6x7.5", "9.0": "6.6x9.0"}

for span_key, df in dfs.items():
    span_label = span_map[span_key]
    csv_fname = f"summary_assemblies_ranked_{span_key}.csv"
    for typology, floor_type, col_type in TYPOLOGIES:
        try:
            col_family_exact = get_col_family_filter(df, col_type)
            best = get_best_row(df, floor_type, col_family_exact)
            rd = build_row_data(best, span_label, typology, csv_fname)
            data[(span_label, typology)] = rd
            col_info = col_family_exact if col_family_exact else "any"
            print(f"  {span_label} {typology}: rank={rd['rank']}, excl_lat={rd['excl_lat_carbon']:.2f} kgCO2e/m² (col={col_info})")
        except ValueError as e:
            print(f"  WARNING: {e}")

print()

# ── Copy source to destination ────────────────────────────────────────────────

print(f"Copying {SRC} -> {DST}")
shutil.copy2(SRC, DST)

# ── Load workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(DST, data_only=False)


# ── 1. Update Assembly Audit sheet ───────────────────────────────────────────

print("\nUpdating Assembly Audit sheet...")
ws_audit = wb["Assembly Audit"]

# Row order: spans then typologies within each span
row_order = []
for span_label in ["6.6x6.6", "6.6x7.5", "6.6x9.0"]:
    for typology, _, _ in TYPOLOGIES:
        row_order.append((span_label, typology))

for i, (span_label, typology) in enumerate(row_order):
    excel_row = i + 2  # row 1 is header
    rd = data.get((span_label, typology))
    if rd is None:
        print(f"  WARNING: No data for {span_label} {typology}")
        continue

    ws_audit.cell(excel_row, 1).value = rd["span"]
    ws_audit.cell(excel_row, 2).value = rd["typology"]
    ws_audit.cell(excel_row, 3).value = rd["rank"]
    ws_audit.cell(excel_row, 4).value = rd["assembly_family"]
    ws_audit.cell(excel_row, 5).value = rd["floor_variant"]
    ws_audit.cell(excel_row, 6).value = rd["beam_variant"]
    ws_audit.cell(excel_row, 7).value = rd["sec_beam_variant"]
    ws_audit.cell(excel_row, 8).value = rd["column_variant"]
    ws_audit.cell(excel_row, 9).value = rd["lateral_variant"]
    ws_audit.cell(excel_row, 10).value = rd["excl_lat_carbon"]
    ws_audit.cell(excel_row, 11).value = rd["total_incl_carbon"]
    ws_audit.cell(excel_row, 12).value = rd["removed_lat_carbon"]
    ws_audit.cell(excel_row, 13).value = rd["concrete_3level_t"]
    ws_audit.cell(excel_row, 14).value = rd["rebar_3level_t"]
    ws_audit.cell(excel_row, 15).value = rd["steel_pt_3level_t"]
    ws_audit.cell(excel_row, 16).value = rd["timber_3level_t"]
    ws_audit.cell(excel_row, 17).value = rd["screed_carbon"]
    ws_audit.cell(excel_row, 18).value = rd["source_csv"]

print(f"  Written {len(row_order)} rows to Assembly Audit")


# ── 2. Update Comparison sheet — My Categories per-m² carbon ─────────────────

print("\nUpdating Comparison sheet (J15:N22 and O3:O10)...")
ws_comp = wb["Comparison"]

# Compute averages across 3 spans for each typology
typology_order = [t for t, _, _ in TYPOLOGIES]  # 8 typologies

for i, typology in enumerate(typology_order):
    comp_row = 15 + i  # rows 15-22

    span_data = [data.get((sl, typology)) for sl in ["6.6x6.6", "6.6x7.5", "6.6x9.0"]]
    span_data = [rd for rd in span_data if rd is not None]

    if not span_data:
        print(f"  WARNING: No span data for {typology}")
        continue

    n = len(span_data)
    avg_concrete_c = sum(rd["concrete_carbon"] for rd in span_data) / n
    avg_rebar_c    = sum(rd["rebar_carbon"] for rd in span_data) / n
    avg_steel_c    = sum(rd["steel_carbon"] for rd in span_data) / n
    avg_timber_c   = sum(rd["timber_carbon"] for rd in span_data) / n
    avg_excl_lat   = sum(rd["excl_lat_carbon"] for rd in span_data) / n

    # Column N = screed + PT + other = total_excl_lateral - concrete - rebar - steel - timber
    avg_other_c = avg_excl_lat - avg_concrete_c - avg_rebar_c - avg_steel_c - avg_timber_c

    ws_comp.cell(comp_row, 10).value = avg_concrete_c  # J
    ws_comp.cell(comp_row, 11).value = avg_rebar_c     # K
    ws_comp.cell(comp_row, 12).value = avg_steel_c     # L
    ws_comp.cell(comp_row, 13).value = avg_timber_c    # M
    ws_comp.cell(comp_row, 14).value = avg_other_c     # N

    print(f"  {typology}: concrete={avg_concrete_c:.2f}, rebar={avg_rebar_c:.2f}, "
          f"steel={avg_steel_c:.2f}, timber={avg_timber_c:.2f}, other={avg_other_c:.2f}")

    # Update O3:O10 (tCO2e for 3 levels) - row = i+3 (typology i -> O row 3+i)
    o_row = 3 + i
    total_tco2e = avg_excl_lat * NUM_LEVELS * GFA_PER_LEVEL / 1000
    ws_comp.cell(o_row, 15).value = total_tco2e  # O column
    print(f"    O{o_row} = {total_tco2e:.2f} tCO2e (3 levels)")


# ── 3. Update span sheets ─────────────────────────────────────────────────────

print("\nUpdating span sheets...")

# Option row locations in each span sheet (row index in Excel)
OPTION_ROWS = {
    "6.6x6.6": {
        "RC Flat Slab": 2,
        "2Way RC Slab": 8,
        "RC Waffle Slab": 16,
        "PT Flat Slab": 24,
        "Composite Steel frame": 31,
        "Steel with Hollow slabs": 41,
        "Steel with CLT Slabs": 52,
        "Glulam with CLT Slabs": 62,
    },
    "6.6x7.5": {
        "RC Flat Slab": 3,
        "2Way RC Slab": 9,
        "RC Waffle Slab": 17,
        "PT Flat Slab": 25,
        "Composite Steel frame": 32,
        "Steel with Hollow slabs": 42,
        "Steel with CLT Slabs": 53,
        "Glulam with CLT Slabs": 63,
    },
    "6.6x9.0": {
        "RC Flat Slab": 3,
        "2Way RC Slab": 9,
        "RC Waffle Slab": 17,
        "PT Flat Slab": 25,
        "Composite Steel frame": 32,
        "Steel with Hollow slabs": 42,
        "Steel with CLT Slabs": 53,
        "Glulam with CLT Slabs": 63,
    },
}

# Column indices (1-based)
COL_R = 18   # R: Option name
COL_T = 20   # T: Weight 1 level (t)
COL_U = 21   # U: Weight 3 levels (t)
COL_V = 22   # V: unit
COL_W = 23   # W: Concrete mass (t, 1 level)
COL_X = 24   # X: Rebar mass (t, 1 level)
COL_Y = 25   # Y: Screed mass (t, 1 level)
COL_Z = 26   # Z: Steel mass (t, 1 level)
COL_AA = 27  # AA: Timber mass (t, 1 level)
COL_AB = 28  # AB: Other = 0
COL_AC = 29  # AC: tCO2eq 1 level
COL_AD = 30  # AD: tCO2eq 3 levels
COL_AE = 31  # AE: Concrete carbon kgCO2e/m²
COL_AF = 32  # AF: Rebar carbon kgCO2e/m²
COL_AG = 33  # AG: Screed carbon kgCO2e/m²
COL_AH = 34  # AH: Steel/PT carbon kgCO2e/m²
COL_AI = 35  # AI: Timber carbon kgCO2e/m²
COL_AJ = 36  # AJ: Other = 0
COL_AK = 37  # AK: kgCO2eq/m² total

for sheet_name, option_rows in OPTION_ROWS.items():
    ws = wb[sheet_name]
    span_label = sheet_name
    print(f"\n  Sheet: {sheet_name}")

    for typology, excel_row in option_rows.items():
        rd = data.get((span_label, typology))
        if rd is None:
            print(f"    WARNING: No data for {typology}")
            continue

        km2 = rd["total_mass_kg_m2"]
        excl_c = rd["excl_lat_carbon"]

        weight_1lv = km2 * GFA_PER_LEVEL / 1000
        weight_3lv = km2 * NUM_LEVELS * GFA_PER_LEVEL / 1000
        tco2_1lv = excl_c * GFA_PER_LEVEL / 1000
        tco2_3lv = excl_c * NUM_LEVELS * GFA_PER_LEVEL / 1000

        # Get steel+PT carbon combined for AH
        steel_pt_carbon = (compute_material_carbon(
            {**dict(zip(
                [f"{c}_carbon_structural_steel_per_m2" for c in ["floor","beam","sec_beam","column"]] +
                [f"{c}_carbon_pt_per_m2" for c in ["floor","beam","sec_beam","column"]],
                [0]*8
            ))}, "x")  # placeholder - compute directly from rd
        )
        # Actually compute from rd fields directly
        steel_c = rd["steel_carbon"]
        # Need PT carbon - compute from data dict
        # We stored steel_carbon but not PT carbon separately, let's compute other = excl - concrete - rebar - steel - timber
        pt_other_c = rd["excl_lat_carbon"] - rd["concrete_carbon"] - rd["rebar_carbon"] - rd["steel_carbon"] - rd["timber_carbon"]
        # But we want steel+PT for AH specifically (not screed)
        # pt_other_c includes screed too
        steel_pt_c = steel_c + (pt_other_c - rd["screed_carbon"])

        ws.cell(excel_row, COL_T).value = weight_1lv
        ws.cell(excel_row, COL_U).value = weight_3lv
        ws.cell(excel_row, COL_V).value = "t"
        ws.cell(excel_row, COL_W).value = rd["concrete_mass_kg_m2"] * GFA_PER_LEVEL / 1000
        ws.cell(excel_row, COL_X).value = rd["rebar_mass_kg_m2"] * GFA_PER_LEVEL / 1000
        ws.cell(excel_row, COL_Y).value = rd["screed_mass_kg_m2"] * GFA_PER_LEVEL / 1000
        ws.cell(excel_row, COL_Z).value = rd["steel_pt_mass_kg_m2"] * GFA_PER_LEVEL / 1000
        ws.cell(excel_row, COL_AA).value = rd["timber_mass_kg_m2"] * GFA_PER_LEVEL / 1000
        ws.cell(excel_row, COL_AB).value = 0
        ws.cell(excel_row, COL_AC).value = tco2_1lv
        ws.cell(excel_row, COL_AD).value = tco2_3lv
        ws.cell(excel_row, COL_AE).value = rd["concrete_carbon"]
        ws.cell(excel_row, COL_AF).value = rd["rebar_carbon"]
        ws.cell(excel_row, COL_AG).value = rd["screed_carbon"]
        ws.cell(excel_row, COL_AH).value = steel_pt_c
        ws.cell(excel_row, COL_AI).value = rd["timber_carbon"]
        ws.cell(excel_row, COL_AJ).value = 0
        ws.cell(excel_row, COL_AK).value = excl_c

        print(f"    {typology} (row {excel_row}): W1={weight_1lv:.1f}t, CO2_1lv={tco2_1lv:.1f}t, kgCO2/m²={excl_c:.2f}")


# ── 4. Fix formula bugs in 6.6x7.5 and 6.6x9.0 ──────────────────────────────

print("\nFixing formula bugs in AN/AO/AP/AQ columns...")

def fix_offbyone_formulas(ws, sheet_name):
    """Fix off-by-one row reference errors in AN/AO/AP/AQ columns."""
    AN, AO, AP, AQ = 40, 41, 42, 43
    fixes = 0

    for row in range(2, ws.max_row + 1):
        for col in [AN, AO, AP, AQ]:
            cell = ws.cell(row, col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue

            # Check for off-by-one: formula references row+1 or row-1 consistently
            # Pattern: look for row numbers in formula that don't match current row
            # E.g., in row 33: =T32-C32 should be =T33-C33

            original = val

            # Find all row references in formula
            # Replace references like C32, T33, AC34, J34 etc.
            # Strategy: find all cell references and check if they should be current row

            # For 6.6x7.5: the bug pattern seems systematic - formulas reference row-1
            # For 6.6x9.0: similar pattern

            # Let's detect by checking if ALL numeric row refs in formula are consistently off
            nums = re.findall(r'[A-Z]{1,2}(\d+)', val)
            if not nums:
                continue

            nums_int = [int(n) for n in nums]
            # If all references are to row-1 (i.e., all are row-1)
            if all(n == row - 1 for n in nums_int):
                new_val = re.sub(r'([A-Z]{1,2})(\d+)', lambda m: m.group(1) + str(row), val)
                if new_val != original:
                    cell.value = new_val
                    fixes += 1
            # If all references are to row+1
            elif all(n == row + 1 for n in nums_int):
                new_val = re.sub(r'([A-Z]{1,2})(\d+)', lambda m: m.group(1) + str(row), val)
                if new_val != original:
                    cell.value = new_val
                    fixes += 1

    # Special fix for 6.6x9.0: AQ formulas with J654-J669 -> J54-J69
    if sheet_name == "6.6x9.0":
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, AQ)
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue
            # Replace J6xx references with J(xx-600)
            def fix_j6xx(m):
                ref_row = int(m.group(1))
                if 600 <= ref_row <= 699:
                    return f"J{ref_row - 600}"
                return m.group(0)
            new_val = re.sub(r'J(\d+)', fix_j6xx, val)
            if new_val != val:
                cell.value = new_val
                fixes += 1
                print(f"    {sheet_name} AQ{row}: {val!r} -> {new_val!r}")

    return fixes


for sheet_name in ["6.6x7.5", "6.6x9.0"]:
    ws = wb[sheet_name]
    n_fixes = fix_offbyone_formulas(ws, sheet_name)
    print(f"  {sheet_name}: {n_fixes} formula fixes applied")

    # Also fix the specific patterns found in inspection
    # For 6.6x7.5 rows 33-38 (Composite section) and rows 43-49 (Hollowcore) and 55-62 (CLT+Steel)
    # For 6.6x9.0 rows 33-38, 43-49, and the #REF! errors

    # Let's do a second pass for the specific off-by-one in certain rows
    AN, AO, AP, AQ = 40, 41, 42, 43

    # Re-check and fix any remaining issues where individual refs are off
    extra_fixes = 0
    for row in range(2, ws.max_row + 1):
        for col in [AN, AO, AP, AQ]:
            cell = ws.cell(row, col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue

            # Replace any cell reference with wrong row number -> correct row number
            # Heuristic: if formula contains refs only off by 1 (mix of row and row-1 or row+1)
            refs = re.findall(r'([A-Z]{1,2})(\d+)', val)
            if not refs:
                continue

            row_nums = [int(r) for _, r in refs]
            # Check if any refs are not == current row
            if any(r != row for r in row_nums):
                # Only fix if refs are row-1 or row+1 (clear typo)
                corrected = val
                for ref_col, ref_row_str in refs:
                    ref_row_int = int(ref_row_str)
                    if abs(ref_row_int - row) == 1:
                        # Replace this specific reference
                        old_ref = ref_col + ref_row_str
                        new_ref = ref_col + str(row)
                        corrected = corrected.replace(old_ref, new_ref, 1)
                if corrected != val:
                    cell.value = corrected
                    extra_fixes += 1

    print(f"  {sheet_name}: {extra_fixes} additional formula fixes applied")


# ── 5. Fix #REF! errors in 6.6x9.0 ──────────────────────────────────────────

print("\nFixing #REF! errors in 6.6x9.0...")
ws90 = wb["6.6x9.0"]
AN, AO, AP, AQ = 40, 41, 42, 43
ref_fixes = 0

for row in range(2, ws90.max_row + 1):
    for col in [AN, AO, AP, AQ]:
        cell = ws90.cell(row, col)
        val = cell.value
        if isinstance(val, str) and "#REF!" in val:
            # Replace #REF! with the correct column reference
            # Based on the pattern: AN58='=T58-#REF!' should be '=T58-C58'
            # AO58='=AN58/#REF!' should be '=AN58/C58'
            # AP58='=(AC58-#REF!)' should be '=(AC58-J58)'
            # Determine what was intended based on col
            if col == AN:  # AN: =Trow-Crow
                new_val = f"=T{row}-C{row}"
            elif col == AO:  # AO: =ANrow/Crow
                new_val = f"=AN{row}/C{row}"
            elif col == AP:  # AP: =(ACrow-Jrow)
                new_val = f"=(AC{row}-J{row})"
            elif col == AQ:  # AQ: =AProw/Jrow
                new_val = f"=AP{row}/J{row}"
            else:
                continue
            cell.value = new_val
            ref_fixes += 1
            print(f"  Fixed #REF! in {cell.column_letter}{row}: {repr(val)[:40]} -> {new_val}")

print(f"  Total #REF! fixes: {ref_fixes}")


# ── Save ──────────────────────────────────────────────────────────────────────

print(f"\nSaving to {DST}...")
wb.save(DST)
print("Done!")

# ── Summary of changes ────────────────────────────────────────────────────────

print("\n" + "="*70)
print("SUMMARY OF UPDATES")
print("="*70)
print(f"\nSource CSVs: 3 EDCA runs (6.6, 7.5, 9.0)")
print(f"Output: {DST}")
print()
print("Assembly Audit (rows 2-25):")
for (span_label, typology), rd in data.items():
    print(f"  {span_label} {typology}: rank={rd['rank']}, excl_lat={rd['excl_lat_carbon']:.2f} kgCO2e/m²")
print()
print("Comparison sheet My Categories (J15:N22) - averages across 3 spans:")
for i, (typology, _, _) in enumerate(TYPOLOGIES):
    span_data = [data.get((sl, typology)) for sl in ["6.6x6.6", "6.6x7.5", "6.6x9.0"]]
    span_data = [rd for rd in span_data if rd is not None]
    if span_data:
        n = len(span_data)
        avg_exc = sum(rd["excl_lat_carbon"] for rd in span_data) / n
        print(f"  {typology}: avg excl_lat={avg_exc:.2f} kgCO2e/m²")
print()
print("Formula fixes applied to 6.6x7.5 and 6.6x9.0 AN/AO/AP/AQ columns.")
print("="*70)
