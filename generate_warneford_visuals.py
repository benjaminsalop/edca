#!/usr/bin/env python3
"""Generate report-ready Warneford comparison charts and cleaned plot data."""

from __future__ import annotations

import argparse
import math
import re
import textwrap
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


SPAN_RE = re.compile(r"^\d+(?:\.\d+)?x\d+(?:\.\d+)?$")
MATERIALS = ["Concrete", "Rebar", "Screed", "Steel", "Wood", "Other"]
REF_MATERIAL_COLS = {"Concrete": 11, "Rebar": 12, "Steel": 13, "Wood": 14, "Other": 15}
TOOL_MATERIAL_COLS = {
    "Concrete": 31,
    "Rebar": 32,
    "Screed": 33,
    "Steel": 34,
    "Wood": 35,
    "Other": 36,
}
COLORS = {
    "Reference": "#355C7D",
    "Tool": "#F67280",
    "Concrete": "#8C8C8C",
    "Rebar": "#C44E52",
    "Screed": "#DDCC77",
    "Steel": "#4C72B0",
    "Wood": "#55A868",
    "Other": "#8172B2",
}


def safe_float(value: Any) -> float:
    if value is None:
        return np.nan
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("#") or stripped.lower() in {"nan", "n/a", "na"}:
            return np.nan
        value = stripped.replace(",", "")
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sheet_to_rows(ws) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def cell(row: list[Any], one_based_col: int) -> Any:
    idx = one_based_col - 1
    return row[idx] if idx < len(row) else None


def normalize_span(span: str) -> str:
    return span.strip()


def span_slug(span: str) -> str:
    return span.replace(".", "_")


def normalize_assembly(raw_name: str, span: str) -> str:
    name = re.sub(re.escape(span), "", raw_name, flags=re.IGNORECASE)
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    mapping = {
        "RC Flat Slab": "RC flat slab",
        "2Way RC Slab": "Two-way RC slab",
        "RC Waffle Slab": "RC waffle slab",
        "PT Flat slab": "PT flat slab",
        "Composite Steel frame": "Composite steel frame",
        "Steel with Hollow slabs": "Steel with hollowcore",
        "Steel with CLT Slabs": "Steel with CLT",
        "Glulam with CLT Slabs": "Glulam with CLT",
    }
    return mapping.get(name, name)


def structural_family(assembly: str) -> str:
    low = assembly.lower()
    if "glulam" in low or ("clt" in low and "steel" not in low):
        return "Timber"
    if "clt" in low and "steel" in low:
        return "Hybrid"
    if "steel" in low or "composite" in low or "hollowcore" in low:
        return "Steel"
    if "rc" in low or "pt" in low or "concrete" in low:
        return "Concrete"
    return "Other"


def component_material(component: str) -> str:
    low = component.lower()
    if "rebar" in low or "tendon" in low:
        return "Rebar"
    if "steel" in low or "beam" in low or "deck" in low:
        return "Steel"
    if "clt" in low or "glulam" in low or "wood" in low or "timber" in low:
        return "Wood"
    if "screed" in low:
        return "Screed"
    if "slab" in low or "column" in low or "concrete" in low:
        return "Concrete"
    return "Other"


def fmt_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def detect_header_row(rows: list[list[Any]]) -> int | None:
    for idx, row in enumerate(rows, start=1):
        values = [clean_text(v).lower() for v in row]
        if "option" in values and any("kgco2eq" in v.replace("\n", "").lower() for v in values):
            return idx
    return None


def inspect_workbook(path: str | Path) -> dict[str, Any]:
    """Inspect workbook sheets, columns, units, row counts, and odd values."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    summary: dict[str, Any] = {
        "path": str(path),
        "sheet_names": wb.sheetnames,
        "sheets": {},
        "assumptions": [
            "Span tabs are detected by names matching '#.#x#.#'.",
            "Assembly rows are detected where column A contains the span name and a numeric total weight/carbon value.",
            "Carbon deviations are tool kgCO2e/m2 minus reference kgCO2e/m2.",
            "Mass deviations by material are inferred from component-level weight differences grouped by component names.",
            "Screed is preserved as its own material category where present.",
        ],
    }
    for ws in wb.worksheets:
        rows = sheet_to_rows(ws)
        nonempty_rows = [r for r in rows if any(v is not None for v in r)]
        header_row = detect_header_row(rows)
        detected_columns = []
        detected_units = set()
        if header_row:
            header = rows[header_row - 1]
            for col_idx, value in enumerate(header, start=1):
                if value is not None:
                    text = clean_text(value).replace("\n", " ")
                    detected_columns.append(f"{col_idx}: {text}")
                    if "kgCO2eq" in text or "kgCO2e" in text:
                        detected_units.add("kgCO2e/m2")
                    if "tCO2" in text:
                        detected_units.add("tCO2e")
                    if "Weight" in text:
                        detected_units.add("t")
        suspicious = []
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                text = clean_text(value)
                if text.startswith("#"):
                    suspicious.append(f"{ws.title}!R{r_idx}C{c_idx}: formula/error value {text}")
                if isinstance(value, (int, float)) and c_idx in {16, 37, 41, 43} and abs(float(value)) > 10_000:
                    suspicious.append(f"{ws.title}!R{r_idx}C{c_idx}: unusually large value {value}")
        summary["sheets"][ws.title] = {
            "max_rows": ws.max_row,
            "max_columns": ws.max_column,
            "nonempty_row_count": len(nonempty_rows),
            "header_row": header_row,
            "detected_columns": detected_columns,
            "detected_units": sorted(detected_units),
            "suspicious_values": suspicious[:200],
            "suspicious_count": len(suspicious),
        }
    return summary


def _parse_span_sheet(span: str, rows: list[list[Any]], notes: list[dict[str, Any]]) -> dict[str, pd.DataFrame]:
    assembly_indices = []
    for idx, row in enumerate(rows):
        option = clean_text(cell(row, 1))
        if span in option and not option.lower().startswith("structural"):
            if not np.isnan(safe_float(cell(row, 3))) or not np.isnan(safe_float(cell(row, 16))):
                assembly_indices.append(idx)

    totals = []
    carbon_material_rows = []
    mass_material_rows = []
    components = []

    for pos, start in enumerate(assembly_indices):
        end = assembly_indices[pos + 1] if pos + 1 < len(assembly_indices) else len(rows)
        row = rows[start]
        raw_assembly = clean_text(cell(row, 1))
        assembly = normalize_assembly(raw_assembly, span)
        family = structural_family(assembly)
        ref_mass_t = safe_float(cell(row, 3))
        tool_mass_t = safe_float(cell(row, 20))
        ref_carbon_t = safe_float(cell(row, 10))
        tool_carbon_t = safe_float(cell(row, 29))
        ref_carbon = safe_float(cell(row, 16))
        tool_carbon = safe_float(cell(row, 37))
        area_m2 = np.nan
        if ref_carbon_t > 0 and ref_carbon > 0:
            area_m2 = 1000 * ref_carbon_t / ref_carbon
        elif tool_carbon_t > 0 and tool_carbon > 0:
            area_m2 = 1000 * tool_carbon_t / tool_carbon
            notes.append(
                {
                    "sheet": span,
                    "assembly": assembly,
                    "issue": "Area inferred from tool carbon because reference area inputs were incomplete.",
                }
            )
        if np.isnan(area_m2):
            notes.append({"sheet": span, "assembly": assembly, "issue": "Could not infer area for kg/m2 mass conversion."})

        carbon_delta = tool_carbon - ref_carbon if not (np.isnan(tool_carbon) or np.isnan(ref_carbon)) else np.nan
        mass_delta_t = tool_mass_t - ref_mass_t if not (np.isnan(tool_mass_t) or np.isnan(ref_mass_t)) else np.nan
        mass_delta_kg_m2 = mass_delta_t * 1000 / area_m2 if not (np.isnan(mass_delta_t) or np.isnan(area_m2) or area_m2 == 0) else np.nan
        carbon_pct = 100 * carbon_delta / ref_carbon if ref_carbon and not np.isnan(carbon_delta) else np.nan
        mass_pct = 100 * mass_delta_t / ref_mass_t if ref_mass_t and not np.isnan(mass_delta_t) else np.nan
        my_option = clean_text(cell(row, 39))
        totals.append(
            {
                "span": span,
                "assembly": assembly,
                "raw_reference_option": raw_assembly,
                "tool_option": my_option,
                "family": family,
                "area_m2": area_m2,
                "reference_mass_t": ref_mass_t,
                "tool_mass_t": tool_mass_t,
                "mass_deviation_t": mass_delta_t,
                "mass_deviation_kg_m2": mass_delta_kg_m2,
                "mass_deviation_pct": mass_pct,
                "reference_carbon_kgCO2e_m2": ref_carbon,
                "tool_carbon_kgCO2e_m2": tool_carbon,
                "carbon_deviation_kgCO2e_m2": carbon_delta,
                "carbon_deviation_pct": carbon_pct,
            }
        )

        for material in MATERIALS:
            ref_val = safe_float(cell(row, REF_MATERIAL_COLS.get(material, -1))) if material in REF_MATERIAL_COLS else 0.0
            tool_val = safe_float(cell(row, TOOL_MATERIAL_COLS.get(material, -1))) if material in TOOL_MATERIAL_COLS else 0.0
            ref_per_m2 = ref_val * 1000 / area_m2 if not (np.isnan(ref_val) or np.isnan(area_m2) or area_m2 == 0) else np.nan
            tool_per_m2 = tool_val if not np.isnan(tool_val) else 0.0
            if material == "Screed":
                ref_per_m2 = 0.0
            delta = tool_per_m2 - (0.0 if np.isnan(ref_per_m2) else ref_per_m2)
            carbon_material_rows.append(
                {
                    "span": span,
                    "assembly": assembly,
                    "family": family,
                    "material": material,
                    "reference_carbon_kgCO2e_m2": ref_per_m2,
                    "tool_carbon_kgCO2e_m2": tool_per_m2,
                    "carbon_deviation_kgCO2e_m2": delta,
                }
            )

        material_mass_delta = defaultdict(float)
        material_mass_ref = defaultdict(float)
        material_mass_tool = defaultdict(float)
        for comp_row in rows[start + 1 : end]:
            ref_comp = clean_text(cell(comp_row, 2))
            tool_comp = clean_text(cell(comp_row, 19))
            if not ref_comp and not tool_comp:
                continue
            component = ref_comp or tool_comp
            if component.lower().startswith("assumed"):
                continue
            material = component_material(component)
            ref_w = safe_float(cell(comp_row, 3))
            tool_w = safe_float(cell(comp_row, 20))
            if np.isnan(ref_w):
                ref_w = 0.0
            if np.isnan(tool_w):
                tool_w = 0.0
            delta_t = tool_w - ref_w
            delta_kg_m2 = delta_t * 1000 / area_m2 if not (np.isnan(area_m2) or area_m2 == 0) else np.nan
            material_mass_delta[material] += 0.0 if np.isnan(delta_kg_m2) else delta_kg_m2
            material_mass_ref[material] += ref_w * 1000 / area_m2 if not (np.isnan(area_m2) or area_m2 == 0) else 0.0
            material_mass_tool[material] += tool_w * 1000 / area_m2 if not (np.isnan(area_m2) or area_m2 == 0) else 0.0
            components.append(
                {
                    "span": span,
                    "assembly": assembly,
                    "component": component,
                    "material": material,
                    "reference_mass_t": ref_w,
                    "tool_mass_t": tool_w,
                    "mass_deviation_t": delta_t,
                    "mass_deviation_kg_m2": delta_kg_m2,
                    "reference_component": ref_comp,
                    "tool_component": tool_comp,
                }
            )
        for material in MATERIALS:
            mass_material_rows.append(
                {
                    "span": span,
                    "assembly": assembly,
                    "family": family,
                    "material": material,
                    "reference_mass_kg_m2": material_mass_ref[material],
                    "tool_mass_kg_m2": material_mass_tool[material],
                    "mass_deviation_kg_m2": material_mass_delta[material],
                }
            )

        if not my_option or my_option.lower() in {"n/a", "na"}:
            notes.append({"sheet": span, "assembly": assembly, "issue": "Tool option is blank or n/a on assembly total row."})
        if np.isnan(ref_carbon) or np.isnan(tool_carbon):
            notes.append({"sheet": span, "assembly": assembly, "issue": "Missing reference or tool carbon total."})
        if not np.isnan(ref_carbon) and ref_carbon == 0:
            notes.append({"sheet": span, "assembly": assembly, "issue": "Reference carbon is zero; percentage error excluded."})
        if np.isnan(ref_mass_t) or np.isnan(tool_mass_t):
            notes.append({"sheet": span, "assembly": assembly, "issue": "Missing reference or tool mass total."})

    return {
        "totals": pd.DataFrame(totals),
        "carbon_material": pd.DataFrame(carbon_material_rows),
        "mass_material": pd.DataFrame(mass_material_rows),
        "components": pd.DataFrame(components),
    }


def load_and_clean_data(path: str | Path) -> dict[str, pd.DataFrame]:
    """Read relevant sheets and return cleaned plotting tables."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    notes: list[dict[str, Any]] = []
    parsed = []
    for sheet in wb.sheetnames:
        if not SPAN_RE.fullmatch(sheet):
            continue
        parsed.append(_parse_span_sheet(normalize_span(sheet), sheet_to_rows(wb[sheet]), notes))

    if not parsed:
        raise ValueError("No span sheets detected. Expected sheet names like 6.6x6.6.")

    totals = pd.concat([p["totals"] for p in parsed], ignore_index=True)
    carbon_material = pd.concat([p["carbon_material"] for p in parsed], ignore_index=True)
    mass_material = pd.concat([p["mass_material"] for p in parsed], ignore_index=True)
    components = pd.concat([p["components"] for p in parsed], ignore_index=True)

    carbon_comparison = totals[
        [
            "span",
            "assembly",
            "family",
            "reference_carbon_kgCO2e_m2",
            "tool_carbon_kgCO2e_m2",
            "carbon_deviation_kgCO2e_m2",
            "carbon_deviation_pct",
        ]
    ].copy()
    mass_comparison = totals[
        [
            "span",
            "assembly",
            "family",
            "reference_mass_t",
            "tool_mass_t",
            "mass_deviation_t",
            "mass_deviation_kg_m2",
            "mass_deviation_pct",
        ]
    ].copy()
    scatter = totals[
        [
            "span",
            "assembly",
            "family",
            "mass_deviation_pct",
            "carbon_deviation_pct",
            "reference_mass_t",
            "reference_carbon_kgCO2e_m2",
        ]
    ].copy()
    before = len(scatter)
    scatter = scatter.replace([np.inf, -np.inf], np.nan)
    dropped = scatter[scatter[["mass_deviation_pct", "carbon_deviation_pct"]].isna().any(axis=1)]
    for _, r in dropped.iterrows():
        notes.append(
            {
                "sheet": r["span"],
                "assembly": r["assembly"],
                "issue": "Excluded from scatter because reference value was zero/missing or percentage deviation was unavailable.",
            }
        )
    scatter = scatter.dropna(subset=["mass_deviation_pct", "carbon_deviation_pct"]).reset_index(drop=True)
    notes.append({"sheet": "all", "assembly": "", "issue": f"Scatter retained {len(scatter)} of {before} assembly-span cases."})

    ranks = []
    for span, group in totals.groupby("span", sort=False):
        tmp = group.copy()
        tmp["reference_rank"] = tmp["reference_carbon_kgCO2e_m2"].rank(method="min", ascending=True)
        tmp["tool_rank"] = tmp["tool_carbon_kgCO2e_m2"].rank(method="min", ascending=True)
        ranks.append(tmp[["span", "assembly", "family", "reference_rank", "tool_rank", "reference_carbon_kgCO2e_m2", "tool_carbon_kgCO2e_m2"]])
    rank_comparison = pd.concat(ranks, ignore_index=True)

    confidence = _build_confidence_table(totals, carbon_material, notes)

    return {
        "totals": totals,
        "carbon_comparison_by_span": carbon_comparison,
        "mass_comparison_by_span": mass_comparison,
        "carbon_deviation_by_material": carbon_material,
        "mass_deviation_by_material": mass_material,
        "carbon_mass_error_scatter": scatter,
        "rank_comparison": rank_comparison,
        "material_components": components,
        "confidence_matrix": confidence,
        "notes": pd.DataFrame(notes),
    }


def _build_confidence_table(totals: pd.DataFrame, carbon_material: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for _, r in totals.iterrows():
        mass_abs = abs(r.get("mass_deviation_pct", np.nan))
        carbon_abs = abs(r.get("carbon_deviation_pct", np.nan))
        quantity_match = 0 if np.isnan(mass_abs) else (3 if mass_abs <= 10 else 2 if mass_abs <= 30 else 1)
        carbon_match = 0 if np.isnan(carbon_abs) else (3 if carbon_abs <= 10 else 2 if carbon_abs <= 30 else 1)
        mats = carbon_material[(carbon_material["span"] == r["span"]) & (carbon_material["assembly"] == r["assembly"])]
        tool_option = clean_text(r.get("tool_option", ""))
        has_screed = bool((mats["material"].eq("Screed") & (mats["tool_carbon_kgCO2e_m2"].abs() > 0)).any())
        has_wood = bool((mats["material"].eq("Wood") & ((mats["reference_carbon_kgCO2e_m2"].abs() > 0) | (mats["tool_carbon_kgCO2e_m2"].abs() > 0))).any())
        boundary_match = 1 if has_screed or tool_option.lower() in {"", "n/a", "na"} else 2
        if quantity_match >= 2 and carbon_match >= 2 and boundary_match >= 2:
            boundary_match = 3
        biogenic = 1 if has_wood else 3
        completeness = 3
        key_values = [r.get("reference_mass_t"), r.get("tool_mass_t"), r.get("reference_carbon_kgCO2e_m2"), r.get("tool_carbon_kgCO2e_m2")]
        if any(pd.isna(v) for v in key_values):
            completeness = 1
        if tool_option.lower() in {"", "n/a", "na"}:
            completeness = min(completeness, 2)
        confidence = int(min(quantity_match, carbon_match, boundary_match, biogenic, completeness))
        rows.append(
            {
                "span": r["span"],
                "assembly": r["assembly"],
                "quantity_match": quantity_match,
                "carbon_match": carbon_match,
                "boundary_match": boundary_match,
                "biogenic_carbon_sensitivity": biogenic,
                "data_completeness": completeness,
                "confidence_level": confidence,
            }
        )
    notes.append({"sheet": "confidence_matrix", "assembly": "", "issue": "Scores are rule-based diagnostics: 3 high, 2 medium, 1 low, 0 incomplete/not comparable."})
    return pd.DataFrame(rows)


def _setup_matplotlib():
    import matplotlib.pyplot as plt

    plt.rcParams.update(
        {
            "font.size": 9,
            "axes.titlesize": 12,
            "axes.labelsize": 9,
            "legend.fontsize": 8,
            "xtick.labelsize": 8,
            "ytick.labelsize": 8,
            "figure.dpi": 120,
            "savefig.dpi": 300,
            "axes.spines.top": False,
            "axes.spines.right": False,
        }
    )
    return plt


def save_figure(fig, outdir: str | Path, stem: str) -> None:
    outdir = Path(outdir)
    fig.savefig(outdir / f"{stem}.png", dpi=300, bbox_inches="tight")
    fig.savefig(outdir / f"{stem}.svg", bbox_inches="tight")


def sorted_spans(data: pd.DataFrame) -> list[str]:
    return sorted(data["span"].dropna().unique(), key=lambda s: [float(x) for x in s.split("x")])


def ordered_assemblies(data: pd.DataFrame) -> list[str]:
    preferred = [
        "RC flat slab",
        "Two-way RC slab",
        "RC waffle slab",
        "PT flat slab",
        "Composite steel frame",
        "Steel with hollowcore",
        "Steel with CLT",
        "Glulam with CLT",
    ]
    present = list(data["assembly"].dropna().unique())
    return [a for a in preferred if a in present] + [a for a in present if a not in preferred]


def make_carbon_comparison_by_span(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["carbon_comparison_by_span"]
    spans = sorted_spans(df)
    assemblies = ordered_assemblies(df)
    fig, axes = plt.subplots(1, len(spans), figsize=(15, 5), sharey=True, constrained_layout=True)
    if len(spans) == 1:
        axes = [axes]
    for ax, span in zip(axes, spans):
        g = df[df["span"] == span].set_index("assembly").reindex(assemblies)
        x = np.arange(len(g))
        width = 0.38
        ax.bar(x - width / 2, g["reference_carbon_kgCO2e_m2"], width, label="Warneford/reference", color=COLORS["Reference"])
        ax.bar(x + width / 2, g["tool_carbon_kgCO2e_m2"], width, label="Tool", color=COLORS["Tool"])
        ax.set_title(span)
        ax.set_xticks(x)
        ax.set_xticklabels(g.index, rotation=45, ha="right")
        ax.set_xlabel("Assembly")
        ax.grid(axis="y", alpha=0.25)
    axes[0].set_ylabel("Embodied carbon (kgCO2e/m2)")
    fig.suptitle("Warneford reference vs tool embodied carbon by grid option", y=1.04, fontsize=14)
    handles, labels = axes[-1].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.02), ncol=2, frameon=False)
    save_figure(fig, outdir, "fig01_carbon_reference_vs_tool_by_span")
    plt.close(fig)


def _diverging_stacked(ax, pivot: pd.DataFrame, xlabel: str, title: str) -> None:
    labels = list(pivot.index)
    y = np.arange(len(labels))
    pos_base = np.zeros(len(labels))
    neg_base = np.zeros(len(labels))
    for material in MATERIALS:
        values = pivot[material].fillna(0).to_numpy() if material in pivot else np.zeros(len(labels))
        positive = np.where(values > 0, values, 0)
        negative = np.where(values < 0, values, 0)
        ax.barh(y, positive, left=pos_base, color=COLORS[material], label=material)
        ax.barh(y, negative, left=neg_base, color=COLORS[material])
        pos_base += positive
        neg_base += negative
    ax.axvline(0, color="#222222", linewidth=0.9)
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_xlabel(xlabel)
    ax.set_title(title)
    ax.grid(axis="x", alpha=0.25)


def make_carbon_deviation_charts(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["carbon_deviation_by_material"]
    letters = "abcdefghijklmnopqrstuvwxyz"
    for i, span in enumerate(sorted_spans(df)):
        g = df[df["span"] == span]
        order = ordered_assemblies(g)
        pivot = g.pivot_table(index="assembly", columns="material", values="carbon_deviation_kgCO2e_m2", aggfunc="sum").reindex(order)
        fig, ax = plt.subplots(figsize=(9, 5.8), constrained_layout=True)
        _diverging_stacked(ax, pivot, "Signed carbon deviation (kgCO2e/m2)", f"Carbon deviation by material category, {span}")
        ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
        save_figure(fig, outdir, f"fig02{letters[i]}_carbon_deviation_{span_slug(span)}")
        plt.close(fig)


def make_mass_deviation_charts(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["mass_deviation_by_material"]
    letters = "abcdefghijklmnopqrstuvwxyz"
    for i, span in enumerate(sorted_spans(df)):
        g = df[df["span"] == span]
        order = ordered_assemblies(g)
        pivot = g.pivot_table(index="assembly", columns="material", values="mass_deviation_kg_m2", aggfunc="sum").reindex(order)
        fig, ax = plt.subplots(figsize=(9, 5.8), constrained_layout=True)
        _diverging_stacked(ax, pivot, "Signed mass deviation (kg/m2)", f"Mass deviation by material category, {span}")
        ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
        save_figure(fig, outdir, f"fig03{letters[i]}_mass_deviation_{span_slug(span)}")
        plt.close(fig)


def make_error_scatter(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["carbon_mass_error_scatter"]
    fig, ax = plt.subplots(figsize=(8, 6), constrained_layout=True)
    families = ["Concrete", "Steel", "Hybrid", "Timber", "Other"]
    markers = {"Concrete": "o", "Steel": "s", "Hybrid": "^", "Timber": "D", "Other": "X"}
    palette = {"Concrete": "#4C72B0", "Steel": "#C44E52", "Hybrid": "#8172B2", "Timber": "#55A868", "Other": "#8C8C8C"}
    for fam in families:
        g = df[df["family"] == fam]
        if g.empty:
            continue
        ax.scatter(g["mass_deviation_pct"], g["carbon_deviation_pct"], s=52, marker=markers[fam], color=palette[fam], label=fam, alpha=0.85)
    ax.axhline(0, color="#222222", linewidth=0.9)
    ax.axvline(0, color="#222222", linewidth=0.9)
    ax.set_xlabel("Mass deviation (%)")
    ax.set_ylabel("Embodied-carbon deviation (%)")
    ax.set_title("Relationship between mass deviation and embodied-carbon deviation")
    ax.grid(alpha=0.25)
    if not df.empty:
        outliers = df.assign(score=df["mass_deviation_pct"].abs() + df["carbon_deviation_pct"].abs()).nlargest(min(5, len(df)), "score")
        for _, r in outliers.iterrows():
            ax.annotate(f"{r['assembly']}\n{r['span']}", (r["mass_deviation_pct"], r["carbon_deviation_pct"]), xytext=(5, 5), textcoords="offset points", fontsize=7)
    ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
    save_figure(fig, outdir, "fig04_carbon_error_vs_mass_error")
    plt.close(fig)


def make_carbon_vs_span(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["carbon_comparison_by_span"]
    spans = sorted_spans(df)
    selected = [
        "RC flat slab",
        "Two-way RC slab",
        "RC waffle slab",
        "PT flat slab",
        "Composite steel frame",
        "Steel with hollowcore",
        "Steel with CLT",
        "Glulam with CLT",
    ]
    complete_tool = []
    for assembly in selected:
        g = df[df["assembly"] == assembly]
        if len(g.dropna(subset=["tool_carbon_kgCO2e_m2"])) >= 2:
            complete_tool.append(assembly)

    fig, ax = plt.subplots(figsize=(8.5, 5.5), constrained_layout=True)
    for assembly in complete_tool:
        g = df[df["assembly"] == assembly].set_index("span").reindex(spans)
        ax.plot(spans, g["tool_carbon_kgCO2e_m2"], marker="o", linewidth=1.8, label=assembly)
    ax.set_title("Tool carbon intensity across grid options")
    ax.set_xlabel("Grid option / span")
    ax.set_ylabel("Embodied carbon (kgCO2e/m2)")
    ax.grid(alpha=0.25)
    ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
    save_figure(fig, outdir, "fig05a_tool_carbon_vs_span")
    plt.close(fig)

    compare = []
    for assembly in selected:
        g = df[df["assembly"] == assembly]
        if len(g.dropna(subset=["reference_carbon_kgCO2e_m2", "tool_carbon_kgCO2e_m2"])) >= 2:
            compare.append(assembly)
    compare = compare[:6]
    fig, ax = plt.subplots(figsize=(9, 5.5), constrained_layout=True)
    cmap = plt.get_cmap("tab10")
    for idx, assembly in enumerate(compare):
        g = df[df["assembly"] == assembly].set_index("span").reindex(spans)
        color = cmap(idx)
        ax.plot(spans, g["reference_carbon_kgCO2e_m2"], marker="o", linewidth=1.8, color=color, label=f"{assembly} reference")
        ax.plot(spans, g["tool_carbon_kgCO2e_m2"], marker="o", linestyle="--", linewidth=1.8, color=color, label=f"{assembly} tool")
    ax.set_title("Reference and tool carbon intensity across grid options")
    ax.set_xlabel("Grid option / span")
    ax.set_ylabel("Embodied carbon (kgCO2e/m2)")
    ax.grid(alpha=0.25)
    ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
    save_figure(fig, outdir, "fig05b_reference_tool_carbon_vs_span_selected")
    plt.close(fig)


def make_rank_comparison(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["rank_comparison"]
    spans = sorted_spans(df)
    fig, axes = plt.subplots(1, len(spans), figsize=(13, 5.8), sharey=True, constrained_layout=True)
    if len(spans) == 1:
        axes = [axes]
    for ax, span in zip(axes, spans):
        g = df[df["span"] == span].sort_values("reference_rank")
        for _, r in g.iterrows():
            ax.plot([0, 1], [r["reference_rank"], r["tool_rank"]], marker="o", color="#555555", alpha=0.8)
            ax.text(-0.04, r["reference_rank"], r["assembly"], ha="right", va="center", fontsize=7)
            ax.text(1.04, r["tool_rank"], r["assembly"], ha="left", va="center", fontsize=7)
        ax.set_xlim(-0.45, 1.45)
        ax.set_xticks([0, 1])
        ax.set_xticklabels(["Reference", "Tool"])
        ax.set_title(span)
        ax.invert_yaxis()
        ax.grid(axis="y", alpha=0.18)
    axes[0].set_ylabel("Carbon rank (1 = lowest)")
    fig.suptitle("Assembly carbon ranking: reference vs tool", y=1.04, fontsize=14)
    save_figure(fig, outdir, "fig06_rank_comparison")
    plt.close(fig)


def make_material_composition(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["carbon_deviation_by_material"].copy()
    rows = []
    for source, col in [("Reference", "reference_carbon_kgCO2e_m2"), ("Tool", "tool_carbon_kgCO2e_m2")]:
        tmp = df[["span", "assembly", "material", col]].rename(columns={col: "carbon"})
        tmp["source"] = source
        rows.append(tmp)
    long = pd.concat(rows, ignore_index=True)
    long["carbon_for_share"] = long["carbon"].clip(lower=0)
    long["case"] = long["span"] + " | " + long["assembly"]
    case_order = [f"{span} | {assembly}" for span in sorted_spans(df) for assembly in ordered_assemblies(df[df["span"] == span])]
    fig, axes = plt.subplots(1, 2, figsize=(14, 8), sharey=True, constrained_layout=True)
    for ax, source in zip(axes, ["Reference", "Tool"]):
        g = long[long["source"] == source]
        pivot = g.pivot_table(index="case", columns="material", values="carbon_for_share", aggfunc="sum").reindex(case_order).fillna(0)
        denom = pivot.sum(axis=1).replace(0, np.nan)
        share = pivot.divide(denom, axis=0).fillna(0) * 100
        left = np.zeros(len(share))
        y = np.arange(len(share))
        for material in MATERIALS:
            vals = share[material].to_numpy() if material in share else np.zeros(len(share))
            ax.barh(y, vals, left=left, color=COLORS[material], label=material)
            left += vals
        ax.set_title(source)
        ax.set_xlim(0, 100)
        ax.set_xlabel("Contribution to positive embodied carbon (%)")
        ax.grid(axis="x", alpha=0.2)
    axes[0].set_yticks(np.arange(len(case_order)))
    axes[0].set_yticklabels(case_order, fontsize=7)
    axes[0].invert_yaxis()
    handles, labels = axes[-1].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.02), ncol=len(MATERIALS), frameon=False)
    fig.suptitle("Material-category contribution to embodied carbon", y=1.02, fontsize=14)
    save_figure(fig, outdir, "fig07_material_composition_fingerprint")
    plt.close(fig)


def make_confidence_matrix(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    plt = _setup_matplotlib()
    df = data["confidence_matrix"].copy()
    df["case"] = df["span"] + " | " + df["assembly"]
    cols = [
        "quantity_match",
        "carbon_match",
        "boundary_match",
        "biogenic_carbon_sensitivity",
        "data_completeness",
        "confidence_level",
    ]
    labels = [
        "quantity\nmatch",
        "carbon\nmatch",
        "boundary\nmatch",
        "biogenic\nsensitivity",
        "data\ncompleteness",
        "confidence\nlevel",
    ]
    matrix = df[cols].to_numpy(dtype=float)
    fig, ax = plt.subplots(figsize=(9, max(5, 0.35 * len(df))), constrained_layout=True)
    im = ax.imshow(matrix, vmin=0, vmax=3, cmap="RdYlGn", aspect="auto")
    ax.set_xticks(np.arange(len(cols)))
    ax.set_xticklabels(labels)
    ax.set_yticks(np.arange(len(df)))
    ax.set_yticklabels(df["case"], fontsize=7)
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            ax.text(j, i, f"{int(matrix[i, j])}", ha="center", va="center", fontsize=8, color="#111111")
    ax.set_title("Boundary/completeness diagnostic matrix")
    cbar = fig.colorbar(im, ax=ax, shrink=0.65)
    cbar.set_label("Score (0 incomplete, 3 high)")
    save_figure(fig, outdir, "fig08_validation_confidence_matrix")
    plt.close(fig)


def export_cleaned_workbook(data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    outdir = Path(outdir)
    output = outdir / "warneford_cleaned_plot_data.xlsx"
    tabs = [
        "carbon_comparison_by_span",
        "mass_comparison_by_span",
        "carbon_deviation_by_material",
        "mass_deviation_by_material",
        "carbon_mass_error_scatter",
        "rank_comparison",
        "confidence_matrix",
        "notes",
    ]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for tab in tabs:
            df = data.get(tab, pd.DataFrame()).copy()
            sheet_name = tab[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                header = str(col_cells[0].value or "")
                max_len = max([len(header)] + [len(fmt_value(c.value)) for c in col_cells[1: min(len(col_cells), 200)]])
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 36)


def write_workbook_summary(summary: dict[str, Any], data: dict[str, pd.DataFrame], outdir: str | Path) -> None:
    lines = []
    lines.append("Warneford workbook inspection summary")
    lines.append("=" * 39)
    lines.append(f"Source workbook: {summary['path']}")
    lines.append("")
    lines.append("Sheet names:")
    for sheet in summary["sheet_names"]:
        lines.append(f"- {sheet}")
    lines.append("")
    for sheet, info in summary["sheets"].items():
        relevant = sheet == "Comparison" or SPAN_RE.fullmatch(sheet)
        if not relevant:
            continue
        lines.append(f"Sheet: {sheet}")
        lines.append(f"- row count: {info['max_rows']} ({info['nonempty_row_count']} non-empty)")
        lines.append(f"- column count: {info['max_columns']}")
        lines.append(f"- detected header row: {info['header_row']}")
        lines.append(f"- detected units: {', '.join(info['detected_units']) if info['detected_units'] else 'none detected'}")
        lines.append("- detected columns:")
        if info["detected_columns"]:
            for col in info["detected_columns"]:
                lines.append(f"  - {col}")
        else:
            lines.append("  - No conventional option/kgCO2e header row detected.")
        lines.append(f"- suspicious/missing values: {info['suspicious_count']} flagged")
        for value in info["suspicious_values"][:20]:
            lines.append(f"  - {value}")
        lines.append("")
    lines.append("Parsed table counts:")
    for key, df in data.items():
        if isinstance(df, pd.DataFrame):
            lines.append(f"- {key}: {len(df)} rows, {len(df.columns)} columns")
    lines.append("")
    lines.append("Dropped/excluded rows and parsing notes:")
    notes = data.get("notes", pd.DataFrame())
    if not notes.empty:
        for _, r in notes.iterrows():
            lines.append(f"- {r.get('sheet', '')} | {r.get('assembly', '')}: {r.get('issue', '')}")
    else:
        lines.append("- None.")
    lines.append("")
    lines.append("Assumptions made while parsing:")
    for assumption in summary.get("assumptions", []):
        lines.append(f"- {assumption}")
    Path(outdir, "warneford_workbook_summary.txt").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Warneford comparison visualizations.")
    parser.add_argument("--input", required=True, help="Path to Warneford_Comparison.xlsx")
    parser.add_argument("--output", required=True, help="Output folder for figures and cleaned workbook")
    args = parser.parse_args()

    outdir = Path(args.output).expanduser()
    outdir.mkdir(parents=True, exist_ok=True)

    summary = inspect_workbook(args.input)
    print("Workbook sheets:")
    for sheet in summary["sheet_names"]:
        print(f"- {sheet}")

    data = load_and_clean_data(args.input)
    write_workbook_summary(summary, data, outdir)
    export_cleaned_workbook(data, outdir)

    make_carbon_comparison_by_span(data, outdir)
    make_carbon_deviation_charts(data, outdir)
    make_mass_deviation_charts(data, outdir)
    make_error_scatter(data, outdir)
    make_carbon_vs_span(data, outdir)
    make_rank_comparison(data, outdir)
    make_material_composition(data, outdir)
    make_confidence_matrix(data, outdir)

    generated = sorted(p.name for p in outdir.iterdir() if p.is_file())
    print("\nGenerated files:")
    print(textwrap.indent("\n".join(generated), "- "))


if __name__ == "__main__":
    main()
